"""
predict.py — Generación de predicciones por ronda y exportación a Excel

Para partidos de fase de grupos: predice Gana Local / Empate / Gana Visitante.
Para partidos de eliminación directa: predice resultado en 90min + probabilidad
de avance asumiendo penales 50/50 en caso de empate.

Uso:
    python predict.py --fixture data/fixtures/semis.csv --round SF --output reports/
"""

import argparse
import pickle
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sklearn.preprocessing import StandardScaler

DATA_DIR    = Path(__file__).parent.parent / "data"
REPORTS_DIR = Path(__file__).parent.parent / "reports"

CLASSES_ORDER = ["A", "D", "H"]  # Away win, Draw, Home win

# Colores para el Excel
NAVY  = "1F3864"
DG    = "1B5E20"
DR    = "C62828"
LG    = "C8E6C9"
LY    = "FFF9C4"
LR    = "FFCDD2"
GR    = "F2F2F2"


# ── Helpers Excel ─────────────────────────────────────────────────────────────
def _hf(color: str) -> PatternFill:
    return PatternFill("solid", start_color=color)

def _bf(bold=True, color="000000", size=10, italic=False) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _al(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, row: int, labels: list, start_col: int = 2, fill: str = NAVY):
    for j, lab in enumerate(labels):
        c = ws.cell(row=row, column=start_col + j, value=lab)
        c.font = _bf(True, "FFFFFF", 9)
        c.fill = _hf(fill)
        c.alignment = _al()
        c.border = _border()


# ── Ensemble predict ──────────────────────────────────────────────────────────
def ensemble_predict(
    X_pred:  np.ndarray,
    models:  dict,
    scaler:  StandardScaler,
) -> np.ndarray:
    """
    Promedia las probabilidades de los modelos del ensemble.
    Devuelve array (n_matches, 3) con orden [P(A), P(D), P(H)].
    """
    all_probas = []
    for name, (m, needs_scale, needs_enc) in models.items():
        Xu = scaler.transform(X_pred) if needs_scale else X_pred
        p  = m.predict_proba(Xu)
        if needs_enc:
            # XGBoost/LightGBM: clases 0=A, 1=D, 2=H
            all_probas.append(p)
        else:
            mc = list(m.classes_)
            p_ord = np.column_stack([p[:, mc.index(c)] for c in CLASSES_ORDER])
            all_probas.append(p_ord)
    return np.mean(all_probas, axis=0)


# ── Calcular columnas de resultado ────────────────────────────────────────────
def add_prediction_columns(
    df: pd.DataFrame,
    proba: np.ndarray,
    is_knockout: bool = False,
) -> pd.DataFrame:
    """
    Añade columnas de probabilidades y pronóstico al DataFrame.

    Fase de grupos: Gana Local / Empate / Gana Visitante
    Eliminación directa: + % Avanza (penales ~50/50)
    """
    df = df.copy()
    df["P_GL"]  = (proba[:, 2] * 100).round(1)   # H
    df["P_E"]   = (proba[:, 1] * 100).round(1)   # D
    df["P_GV"]  = (proba[:, 0] * 100).round(1)   # A

    def pronostico(r):
        p = {"Gana Local": r["P_GL"], "Empate": r["P_E"], "Gana Visitante": r["P_GV"]}
        return max(p, key=p.get)

    df["Pronostico"] = df.apply(pronostico, axis=1)
    df["Confianza"]  = df[["P_GL", "P_E", "P_GV"]].max(axis=1)

    if is_knockout:
        # Penales ~50/50 si empate en 90min
        df["P_Avanza_Local"]  = ((proba[:, 2] + 0.5 * proba[:, 1]) * 100).round(1)
        df["P_Avanza_Visit"]  = ((proba[:, 0] + 0.5 * proba[:, 1]) * 100).round(1)
        df["Quien_Avanza"]    = df.apply(
            lambda r: r["home_team"] if r["P_Avanza_Local"] >= 50 else r["away_team"],
            axis=1,
        )
        df["Conf_Avance"]     = df[["P_Avanza_Local", "P_Avanza_Visit"]].max(axis=1)
        df["Pronostico"]      = df.apply(
            lambda r: "Gana Local" if r["P_GL"] > r["P_GV"] and r["P_GL"] > r["P_E"]
            else ("Empate (TP)" if r["P_E"] >= r["P_GL"] and r["P_E"] >= r["P_GV"]
                  else "Gana Visitante"),
            axis=1,
        )

    return df


# ── Exportar a Excel ──────────────────────────────────────────────────────────
def export_excel(
    pred_df:    pd.DataFrame,
    output_path: Path,
    round_name:  str = "Predicciones",
    model_info:  str = "",
    is_knockout: bool = False,
) -> None:
    """Genera un archivo Excel formateado con las predicciones."""
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Predicciones"
    ws.sheet_view.showGridLines = False

    ws["B2"] = f"⚽  {round_name} — Copa Mundial FIFA 2026"
    ws["B2"].font = _bf(True, NAVY, 14)

    ws["B3"] = model_info or "Modelo: Ensemble v6 (Logística + RF + RF Calibrado) · 20 features"
    ws["B3"].font = _bf(False, "595959", 9, italic=True)

    if is_knockout:
        ws["B4"] = "★ Sin empate real en eliminación directa. % Avanza = Gana 90min + 50%×Empate (penales ~50/50)"
        ws["B4"].font = _bf(False, DR, 9, italic=True)

    # Cabeceras
    start = 5
    if is_knockout:
        labs = ["Fecha", "Local", "Visitante", "Gana Local 90m", "Empate 90m",
                "Gana Visit. 90m", "% Avanza Local", "% Avanza Visit.",
                "Quién Avanza", "Confianza"]
        cols = ["date", "home_team", "away_team", "P_GL", "P_E", "P_GV",
                "P_Avanza_Local", "P_Avanza_Visit", "Quien_Avanza", "Conf_Avance"]
    else:
        labs = ["Fecha", "Local", "Visitante", "% Gana Local", "% Empate",
                "% Gana Visit.", "Pronóstico", "Confianza"]
        cols = ["date", "home_team", "away_team", "P_GL", "P_E", "P_GV",
                "Pronostico", "Confianza"]

    _header_row(ws, start, labs)

    TEAM_COLORS = {
        "France": "1A237E", "Spain": "7B1FA2", "Argentina": "1B5E20",
        "England": "0D47A1", "Morocco": "006400", "Belgium": "4A148C",
        "Norway": "C62828", "Switzerland": "B71C1C", "Brazil": "006400",
        "Colombia": "C62828", "Portugal": "880E4F", "Netherlands": "E65100",
    }

    for i, row in pred_df.sort_values("date").reset_index(drop=True).iterrows():
        er  = start + 1 + i
        alt = i % 2 == 1

        for j, col in enumerate(cols):
            val = row.get(col, "")
            if col == "date":
                try:
                    val = pd.to_datetime(str(val)).strftime("%d/%m")
                except Exception:
                    pass

            c = ws.cell(row=er, column=2 + j, value=val)
            c.border = _border()

            # Columna de favorito / quién avanza
            winner_col = "Quien_Avanza" if is_knockout else "Pronostico"
            if col == winner_col and is_knockout:
                fc = TEAM_COLORS.get(str(val), NAVY)
                c.font = _bf(True, "FFFFFF", 10)
                c.fill = _hf(fc)
            elif col in ("Confianza", "Conf_Avance"):
                conf_val = float(val) if val else 50
                bg = LG if conf_val >= 65 else (LY if conf_val >= 52 else LR)
                c.font = _bf(True)
                c.fill = _hf(bg)
            else:
                c.font = _bf(False)
                c.fill = _hf(GR if alt else "FFFFFF")
                if col in ["home_team", "away_team", "Pronostico"]:
                    c.alignment = _al("left")
                    continue

            c.alignment = _al()

    # Anchos de columna
    widths = [3, 9, 20, 20, 12, 11, 14, 12, 12, 18, 12]
    for k, w in enumerate(widths[: len(cols) + 1]):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(k + 2)].width = w

    # Leyenda
    leg = start + len(pred_df) + 2
    ws.cell(row=leg, column=2, value="Leyenda confianza:").font = _bf(True, NAVY)
    for jl, (lab, bg) in enumerate([("Alta ≥65%", LG), ("Media 52-64%", LY), ("Baja <52%", LR)]):
        c = ws.cell(row=leg, column=3 + jl, value=lab)
        c.font = _bf(True, "000000", 9)
        c.fill = _hf(bg)
        c.alignment = _al()
        c.border = _border()

    wb.save(output_path)
    print(f"📊 Excel guardado en: {output_path}")


# ── Función principal ─────────────────────────────────────────────────────────
def predict(
    fixture_path:  Path,
    models_path:   Path = DATA_DIR / "trained_models.pkl",
    output_dir:    Path = REPORTS_DIR,
    round_name:    str  = "Predicciones",
    is_knockout:   bool = False,
) -> pd.DataFrame:
    """
    Genera predicciones para un fixture dado usando el ensemble entrenado.

    Args:
        fixture_path:  CSV con columnas [home_team, away_team, date, neutral?, city?]
        models_path:   Ruta al archivo .pkl con el ensemble entrenado
        output_dir:    Directorio donde guardar el Excel
        round_name:    Nombre de la ronda (aparece en el Excel)
        is_knockout:   True si es fase eliminatoria (sin empate real)

    Returns:
        DataFrame con predicciones
    """
    # Cargar modelo
    with open(models_path, "rb") as f:
        artifact = pickle.load(f)

    models        = artifact["models"]
    scaler        = artifact["scaler"]
    good_feats    = artifact["features"]
    state_path    = DATA_DIR / "feature_state.pkl"

    # Cargar estado de features
    with open(state_path, "rb") as f:
        state_data = pickle.load(f)

    from features import build_fixture_features, FeatureState
    from elo import EloEngine

    engine = state_data["engine"]
    state  = state_data["state"]

    # Cargar fixture
    fixture = pd.read_csv(fixture_path)
    if "neutral" not in fixture.columns:
        fixture["neutral"] = True  # knockout siempre es neutral
    if "is_world_cup" not in fixture.columns:
        fixture["is_world_cup"] = True

    # Construir features
    feat_df = build_fixture_features(fixture, engine, state)
    feat_df["is_ko"] = 1 if is_knockout else 0

    # Asegurar que existen todas las features necesarias
    for f in good_feats:
        if f not in feat_df.columns:
            feat_df[f] = 0.0

    X_pred = feat_df[good_feats].values

    # Predicción del ensemble
    proba = ensemble_predict(X_pred, models, scaler)

    # Armar resultado
    pred_df = fixture.copy()
    pred_df = add_prediction_columns(pred_df, proba, is_knockout)

    # Guardar CSV
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path   = output_dir / f"{round_name.replace(' ', '_')}_predicciones.csv"
    excel_path = output_dir / f"{round_name.replace(' ', '_')}_predicciones.xlsx"

    pred_df.to_csv(csv_path, index=False)
    export_excel(pred_df, excel_path, round_name, is_knockout=is_knockout)

    print(f"\n=== PREDICCIONES: {round_name} ===")
    cols_show = (
        ["home_team", "away_team", "P_GL", "P_E", "P_GV", "Quien_Avanza", "Conf_Avance"]
        if is_knockout
        else ["home_team", "away_team", "P_GL", "P_E", "P_GV", "Pronostico", "Confianza"]
    )
    print(pred_df[cols_show].to_string(index=False))
    return pred_df


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Predicciones WC2026 Predictor")
    parser.add_argument("--fixture",  required=True, help="CSV con el fixture de la ronda")
    parser.add_argument("--round",    default="Predicciones", help="Nombre de la ronda")
    parser.add_argument("--knockout", action="store_true",
                        help="Fase eliminatoria (sin empate real)")
    parser.add_argument("--models",   default=str(DATA_DIR / "trained_models.pkl"))
    parser.add_argument("--output",   default=str(REPORTS_DIR))
    args = parser.parse_args()

    predict(
        fixture_path=Path(args.fixture),
        models_path=Path(args.models),
        output_dir=Path(args.output),
        round_name=args.round,
        is_knockout=args.knockout,
    )
